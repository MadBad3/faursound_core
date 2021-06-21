import sys
from df2NVHExpert import df2NVHExpert
import pytest, os, sys, pandas as pd
# sys.path.append((os.path.dirname(os.path.dirname(__file__))))

from openpyxl import load_workbook
from etiltWav import etiltWav

path2stft = r'./test'
path2prediction = r'./test/text'
path2csv = r'./test/csv_folder'
text_file = 'E-Tilt_6202-PC1_2402749_20210201_052215_11ACF9023_750N - Down.txt'
void_text_file = 'Void.txt'
path2label_map_json = os.path.join(path2stft+'/label_map.json')
listOfColumnsNameDf = ['txt_file_name', 'grinding-score', 'grinding-criticalLevel', 'grinding-sorting_score', 'Wave', 'prediction_pic']
listOfColumnsNameDf2 = ['object_id', 'score', 'label', 'criticalLevel', 'sorting_score']
totalLabels = len(etiltWav.LABELS)
grinding_label = 'grinding'

def test_load_label_map_from_json():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	data = obj.load_label_map_from_json(path2label_map_json)
	assert type(data) == dict and len(data) == totalLabels

def test_process_txt_folder_to_df():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	output = obj.process_txt_folder_to_df()
	assert type(output) == pd.DataFrame
	assert output.columns.values.tolist() == listOfColumnsNameDf

def test_process_one_txt_file():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	output = obj._process_one_txt_file(text_file)
	assert type(output) == pd.DataFrame and output.shape == (1, len(listOfColumnsNameDf[:-2]))
	assert output.columns.values.tolist() == listOfColumnsNameDf[:-2]
	with pytest.raises(ValueError, match=r'not txt file passed for _process_one_txt_file'):
		output = obj._process_one_txt_file('text_file.py')

def test_get_different_ext_file_name():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	output = obj._get_different_ext_file_name(text_file)
	assert type(output) == str and output == 'E-Tilt_6202-PC1_2402749_20210201_052215_11ACF9023_750N - Down.jpg'

def test_read_txt_to_df():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	
	stft_file = obj._get_different_ext_file_name(text_file)
	txt_path = os.path.join(obj.path2prediction,text_file)
	stft_path = os.path.join(obj.path2stft,stft_file)
	output = obj._read_txt_to_df(txt_path, stft_path)

	file = open(txt_path, "r")
	nonempty_lines = [line.strip("\n") for line in file if line != "\n"]

	assert type(output) == pd.DataFrame and grinding_label in output.values
	assert output.columns.values.tolist() == listOfColumnsNameDf2[:-1] and output.shape == (len(nonempty_lines)-1, len(listOfColumnsNameDf2[:-1]))

	txt_path = os.path.join(obj.path2stft,'void_text',void_text_file)
	output = obj._read_txt_to_df(txt_path, stft_path)
	assert output == None

def test_only_one_object_per_label():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)

	stft_file = obj._get_different_ext_file_name(text_file)
	txt_path = os.path.join(obj.path2prediction,text_file)
	stft_path = os.path.join(obj.path2stft,stft_file)
	df = obj._read_txt_to_df(txt_path, stft_path)
	
	output = obj._only_one_object_per_label(df)

	assert type(output) == pd.DataFrame
	assert output.shape == (len(listOfColumnsNameDf2),1)
	assert output.columns.values.tolist() == [grinding_label]

def test_flatten_df_2():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)

	stft_file = obj._get_different_ext_file_name(text_file)
	txt_path = os.path.join(obj.path2prediction,text_file)
	stft_path = os.path.join(obj.path2stft,stft_file)
	df = obj._read_txt_to_df(txt_path, stft_path)
	df2 = obj._only_one_object_per_label(df)

	output = obj._flatten_df_2(df2)

	assert type(output) == pd.DataFrame and output.shape == (1, len(obj.items))
	
def test_create_empty_df():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	output = obj._create_empty_df()
	assert type(output) == pd.DataFrame and output.shape == (1, totalLabels*len(obj.items))

def test_labels():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	output = obj.labels
	assert len(output) == totalLabels

def test_items():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	output = obj.items
	assert type(output) == list

def test_add_wav_fn_column():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	example = {'txt_file_name': ['example1','example2']}

	df_merged = pd.DataFrame(example)
	
	listBefore = list(df_merged.shape)
	listBefore[-1] = listBefore[-1]+1
	shapeAfter = tuple(listBefore)
	obj._add_wav_fn_column(df_merged)
	assert df_merged.shape == shapeAfter and any(df_merged.Wave)

def test_add_prediction_pic_fn_column():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	example = {'txt_file_name': ['example1','example2']}

	df_merged = pd.DataFrame(example)
	
	listBefore = list(df_merged.shape)
	listBefore[-1] = listBefore[-1]+1
	shapeAfter = tuple(listBefore)
	obj._add_prediction_pic_fn_column(df_merged)
	assert df_merged.shape == shapeAfter and any(df_merged.prediction_pic)

def test_save_df_to_excel():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	obj.save_df_to_excel()
	os.remove(obj.excel_path)

def test_load_csv_meta():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json)
	obj.load_csv_meta(os.path.join(path2csv+'/sample.csv'))

def test_add_link_to_column():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json, path2csv)
	obj._add_link_to_column('txt_file_name', path2prediction)

	wb = load_workbook(os.path.join(path2csv+'/detection_with_cl.xlsx'))
	ws = wb['Sheet1']
	assert ws.cell(row=2, column=1).hyperlink.target
	file_path = os.path.relpath(ws.cell(row=2, column=1).hyperlink.target)
	assert open(file_path)
	
	df = pd.read_excel(os.path.join(path2csv+'/detection_with_cl.xlsx'))
	df.to_excel(os.path.join(path2csv+'/detection_with_cl.xlsx'), sheet_name='Sheet1', index=False)

def test_add_hyper_links():
	obj = df2NVHExpert(path2prediction, path2stft, path2label_map_json, path2csv)
	obj.add_hyper_links(path2stft, path2stft)
	
	wb = load_workbook(os.path.join(path2csv+'/detection_with_cl.xlsx'))
	ws = wb['Sheet1']
	
	assert ws.cell(row=2, column=1).hyperlink.target
	assert ws.cell(row=2, column=5).hyperlink.target
	assert ws.cell(row=2, column=6).hyperlink.target

	file_path1 = ws.cell(row=2, column=1).hyperlink.target
	file_path2 = ws.cell(row=2, column=5).hyperlink.target
	file_path3 = ws.cell(row=2, column=6).hyperlink.target
	assert open(file_path1) and open(file_path2) and open(file_path3)

	df = pd.read_excel(os.path.join(path2csv+'/detection_with_cl.xlsx'))
	df.to_excel(os.path.join(path2csv+'/detection_with_cl.xlsx'), sheet_name='Sheet1', index=False)