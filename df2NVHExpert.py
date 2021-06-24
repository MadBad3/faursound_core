import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import json
import sys
# sys.path.append((os.path.dirname(os.path.dirname(__file__))))

from criticalLevel import criticalLevel

# this class take prediction result (txt file) and original picture, calculation critical level and export xls for NVH expert analysis

class df2NVHExpert(object):

    def __init__(self, path2prediction:str, path2stft:str, path2label_map_json:str, output_path:str = ''):
        """[init function]

        :param path2prediction: [path to prediction txt files]
        :type path2prediction: str
        :param path2stft: [path to origin stft picture]
        :type path2stft: str
        :param path2label_map_json: [path to label map json file]
        :type path2label_map_json: str
        :param output_path: [output path], defaults to ''
        :type output_path: str, optional
        """

        self.path2prediction = path2prediction
        self.path2stft = path2stft

        if output_path:
            self.output_path = output_path
        else:
            self.output_path = self.path2prediction

        self.category_index = self.load_label_map_from_json(path2label_map_json)

        self.df = self.process_txt_folder_to_df()
        self.excel_path = os.path.join(self.output_path, r'detection_with_cl.xlsx')


    def load_label_map_from_json(self,json_fn:str) -> dict:
        """load label map from json file.

        :param json_fn: [json file name]
        :type json_fn: [str]
        :return: [dict with label map]
        :rtype: dict
        """
        def int_keys(ordered_pairs):
            result = {}
            for key, value in ordered_pairs:
                try:
                    key = int(key)
                except ValueError:
                    pass
                result[key] = value
            return result

        with open(json_fn) as f:
            data = json.load(f,object_pairs_hook=int_keys)
        return data


    def _read_txt_to_df(self, txt_file_path:str, stft_file_path:str) -> pd.DataFrame:
        """[read txt file into df]

        :param txt_file_path: [txt file path]
        :type txt_file_path: str
        :param stft_file_path: [stft file path]
        :type stft_file_path: str
        :return: [dataFrame with information needed]
        :rtype: pd.DataFrame
        """
        cl = criticalLevel()

        df = pd.read_csv(txt_file_path, delimiter = "\n", header = 0) # row0 is name of the file. this avoid empty txt file.
                                                                        #1 object = 1 row. each row have x1,y1,x2,y2,score,label
        temp = []
        label_offset = 0            # no more need this offset, was done already during inference
        num_of_object_output = len(df)

        if num_of_object_output > 0:
            for idx in range(num_of_object_output):
                infor = df.iloc[idx].item().split(',')
                infor[-1] = self.category_index[int(infor[-1]) + label_offset]['name'] #infor[-1] is where label stores, then we decoded it to label name
                cl_value = cl.run(stft_file_path,x1=infor[0],y1=infor[1],x2=infor[2],y2=infor[3],
                                score=infor[4],label=infor[5])
                series = pd.Series(data = {
                            'object_id':f"{idx:02d}",
                            'score':round(float(infor[4]),2),
                            'label':infor[5],
                            'criticalLevel':cl_value,
                            })

                temp.append(series)

            df_1 = pd.DataFrame(temp)

            return df_1
        else:
            return None


    def _only_one_object_per_label(self, df_1:pd.DataFrame) -> pd.DataFrame:
        """keep only highest sorting score per class

        :param df_1: [df_1 with all detection]
        :type df: pd.DataFrame
        :return: [df_2 with max one detection per class]
        :rtype: pd.DataFrame
        """
        def calculate_sorting_score(row) -> float:
            if row['score'] > 0.9:
                sorting_score = row['criticalLevel']
            else :
                sorting_score = row['criticalLevel'] * row['score']
            return sorting_score
        
        df_1['sorting_score'] = df_1.apply(calculate_sorting_score, axis=1)

        df_2 = df_1.loc[df_1.groupby('label')['sorting_score'].idxmax(), :].reset_index(drop= True).T
        df_2.columns = df_2.loc['label']

        return df_2


    def _flatten_df_2(self, df_2:pd.DataFrame) ->pd.DataFrame:
        """flatten df

        :param df_2: [description]
        :type df_2: pd.DataFrame
        :return: [description]
        :rtype: pd.DataFrame
        """
        df_3 = df_2.loc[self.items]

        place_holder = {}
        for label in df_3.columns:
            for item in df_3.index:
                column_name = label + '-' + item
                place_holder[column_name] = df_3.loc[item,label]
        s = pd.Series(place_holder)
        return pd.DataFrame(s).T  


    @property
    def labels(self):
        return [x['name'] for x in self.category_index.values()]


    @property
    def items(self):
        return ['score','criticalLevel','sorting_score']


    def process_txt_folder_to_df(self) -> pd.DataFrame:
        """process txt folder to df

        :return: [data frame for output]
        :rtype: pd.DataFrame
        """

        dfs = []
        for txt_file in os.listdir(self.path2prediction):
            df_3 = self._process_one_txt_file(txt_file)
            dfs.append(df_3)

        df_merged = pd.concat(dfs, ignore_index=True, axis = 0,  join='outer')
        self._add_wav_fn_column(df_merged)
        self._add_prediction_pic_fn_column(df_merged)
        return df_merged


    def _process_one_txt_file(self, txt_file:str) -> pd.DataFrame:
        """preprocess one txt file

        :param txt_file: [txt file path]
        :type txt_file: str
        :raises ValueError: [description]
        :return: [df for one txt file]
        :rtype: pd.DataFrame
        """
        _ , file_extension = os.path.splitext(txt_file)

        if file_extension == '.txt':
            stft_file = self._get_different_ext_file_name(input_fn=txt_file,target_ext='.jpg')

            txt_path = os.path.join(self.path2prediction,txt_file)
            stft_path = os.path.join(self.path2stft,stft_file)

            df_1 = self._read_txt_to_df(txt_path, stft_path)

            if isinstance(df_1, pd.DataFrame):
                df_2 = self._only_one_object_per_label(df_1)
                df_3 = self._flatten_df_2(df_2)
            else :
                df_3 = self._create_empty_df()

            df_3.insert(0, 'txt_file_name', os.path.basename(txt_path), allow_duplicates=False)
        else :
            raise ValueError('not txt file passed for _process_one_txt_file')
        return df_3


    def _create_empty_df(self) -> pd.DataFrame:
        """create a empty df

        :return: [empty df]
        :rtype: pd.DataFrame
        """
        place_holder = {}
        for label in self.labels:
            for item in self.items:
                column_name = label + '-' + item
                place_holder[column_name] = np.nan

        s = pd.Series(place_holder)
        return pd.DataFrame(s).T  


    def _get_different_ext_file_name(self, input_fn:str, target_ext:str ='.jpg') -> str:
        """get different ext file name based on file name

        :param input_fn: [input file name]
        :type input_fn: str
        :param target_ext: [target extension], defaults to '.jpg'
        :type target_ext: str, optional
        :return: [new file name with target extension]
        :rtype: str
        """
        title , ext = os.path.splitext(input_fn)
        return f'{title}{target_ext}'


    def save_df_to_excel(self):
        """save df to excel
        """

        self.df.to_excel(self.excel_path,index = False)
        print(f'save excel to {self.excel_path} -> Done !')


    def return_json_from_one_txt(self,txt_file):

        df3 = self._process_one_txt_file(txt_file = txt_file)
        result = df3.iloc[:,1:].to_json(orient="records")
        return result
        

    def return_json_from_one_txt_full(self,txt_file):

        df3 = self._process_one_txt_file(txt_file = txt_file)
        df3_to_merge = df3.iloc[:,1:]
        df_empty = self._create_empty_df()
        for column in df3_to_merge.columns:
            df_empty[column] = df3_to_merge[column]  # fill it the value to df_empty

        result = df_empty.to_json(orient="records")
        return result


    def _add_wav_fn_column(self, df:pd.DataFrame) -> None:
        """add wav fn column to df

        :param df: [df to add info]
        :type df: pd.DataFrame
        """
        def get_wav_fn_based_on_txt_fn(row):
            title , _ = os.path.splitext(row.txt_file_name)
            return f'{title}.wav'

        df['Wave'] = df.apply(get_wav_fn_based_on_txt_fn, axis = 1)


    def _add_prediction_pic_fn_column(self, df:pd.DataFrame) -> None:
        """add prediction pic fn column

        :param df: [df to add info]
        :type df: pd.DataFrame
        """
        def get_png_fn_based_on_txt_fn(row):
            title , _ = os.path.splitext(row.txt_file_name)
            return f'{title}.png'

        df['prediction_pic'] = df.apply(get_png_fn_based_on_txt_fn, axis = 1)


    def load_csv_meta(self, csv_path:str):
        """load csv meta 

        :param csv_path: [csv file path]
        :type csv_path: str
        """
        df_meta = pd.read_csv(csv_path, header = 0)
        self.df = df_meta.merge(self.df, how= 'inner', on = 'Wave') 


    def _add_link_to_column(self, column_name:str, path2link:str):
        """add hyperlink to excel column

        :param column_name: [column name]
        :type column_name: str
        :param path2link: [path for the hyperlink]
        :type path2link: str
        """

        wb = load_workbook(self.excel_path) 
        ws1 = wb.get_sheet_by_name("Sheet1")

        abspath = os.path.abspath(path2link)

        column_list = self.df.columns.to_list()
        column_index = column_list.index(column_name)
        for i in range(len(self.df)):
            fn = self.df[column_name][i]
            link = os.path.join(abspath, fn)

            ws1.cell(row= i + 2, column = column_index + 1).hyperlink = link    # row = i+2 because first row is column name
            ws1.cell(row= i + 2, column = column_index + 1).style = "Hyperlink"

        wb.save(self.excel_path)


    def add_hyper_links(self,wav_folder:str = '', prediction_pic_folder:str = ''):
        """add hyperlink for excel file. optional for what to add

        :param wav_folder: [wav folder], defaults to ''
        :type wav_folder: str, optional
        :param prediction_pic_folder: [prediction pic folder], defaults to ''
        :type prediction_pic_folder: str, optional
        """
        self._add_link_to_column('txt_file_name', self.path2prediction)

        if wav_folder:
            self._add_link_to_column('Wave', wav_folder)

        if prediction_pic_folder:
            self._add_link_to_column('prediction_pic', prediction_pic_folder)
            
        print(f'add_hyper_links_to_excel {self.excel_path}-> Done')

